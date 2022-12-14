from configParser import Settings
from win32com.client import Dispatch
from win32com.client import gencache
import pythoncom
import pandas as pd
import openpyxl
import numpy as np

class Methods:

    def reportGenerate(self,row,sheet_obj,coll):
        excel_data = []

        for i in range(coll,coll+21):
            excel_data.append(sheet_obj.cell(row=row, column=i).value)


        data = pd.read_csv(f'csv\star{row}.csv',sep=';',decimal=',',encoding = "ISO-8859-1")
        
        list_of_max = []

        for i in data:
            list_of_max.append(data[i].max())

        wb = openpyxl.load_workbook('report/report.xlsx')
        
        sheet = wb.active

        for i in range(len(excel_data)):
            sheet.cell(row = row, column=coll+i).value = excel_data[i]
        
        for i in range(len(list_of_max)):
            sheet.cell(row = row, column=coll+i + len(excel_data)).value = str(list_of_max[i])

        wb.save('report/report.xlsx')  

        header = ['N1','L1','N2','L2','a2','N3','L3','','','','Air blades','Air inlet','СH4','Temperature','S2','NН3','Н2','O2','N2','CO2','H20',
        'Line Probe 8: Direction [-1,0,0] (m)',
        'Line Probe 8: Temperature (K)',
        'Line Probe 8: Direction [-1,0,0] (m)',
        'Line Probe 8: Mass Fraction of Nitrogen Oxide Emission',
        'Line Probe 8: Direction [-1,0,0] (m)',
        'Line Probe 8: Mass Fraction of CO',
        'Line Probe 8: Direction [-1,0,0] (m)',
        'Line Probe 8: Mass Fraction of H2O']

        csv_data =[[i] for i in excel_data+list_of_max]

        for i in range(len(csv_data)):
            for j in range(len(csv_data[i])):
                csv_data[i][j] = str(csv_data[i][j]).replace('.',',')

        df2 = pd.DataFrame(np.array(csv_data).transpose(),
                   columns=header)

        df2.to_csv('report/report.csv',sep=';',mode='a',decimal=',', float_format='%.3f', header=False)



    def numberOfcullums(self,sheet_obj,row_start,coll):
        row_number = 0
        row = row_start
        while True:
            cell_value = sheet_obj.cell(row=row, column=coll).value
            if cell_value == None:
                return row_number
            else:
                row_number = row_number + 1
            
            row = row + 1

    def checkRepeatRow(self,sheet_obj,row,coll):
        settings = Settings()
        start_geometry = []

        for i in range(coll,coll+7):
            start_geometry.append(sheet_obj.cell(row=row, column=i).value)

        for i in reversed(range(int(settings.get_from_settings("excel_start_row")),row)):
            check_geomentry = [] 
            for j in range(coll,coll+7):
                check_geomentry.append(sheet_obj.cell(row=i, column=j).value)
            
            if start_geometry == check_geomentry:
                return i
            
        return 0

    def changeGeom(self,geomPath,sheet_obj,row):
        def get_kompas_api7():
                module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
                api = module.IKompasAPIObject(
                        Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(module.IKompasAPIObject.CLSID,
                                                                                pythoncom.IID_IDispatch))
                const = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants
                return module, api, const
        

        settings = Settings()

        module7, api7, const7 = get_kompas_api7()  # Подключаемся к API7
        app7 = api7.Application  # Получаем основной интерфейс
        app7.Visible = True  # Показываем окно пользователю (если скрыто)
        app7.HideMessage = const7.ksHideMessageNo  # Отвечаем НЕТ на любые вопросы программы
        print(app7.ApplicationName(FullName=True))  # Печатаем название программы

        doc7 = app7.Documents.Open(PathName=geomPath,
                                Visible=True,
                                ReadOnly=True)

        #  Подключим константы API Компас
        kompas6_constants = gencache.EnsureModule("{75C9F5D0-B5B8-4526-8681-9903C567D2ED}", 0, 1, 0).constants
        kompas6_constants_3d = gencache.EnsureModule("{2CAF168C-7961-4B90-9DA2-701419BEEFE3}", 0, 1, 0).constants

        #  Подключим описание интерфейсов API5
        kompas6_api5_module = gencache.EnsureModule("{0422828C-F174-495E-AC5D-D31014DBBE87}", 0, 1, 0)
        kompas_object = kompas6_api5_module.KompasObject(
            Dispatch("Kompas.Application.5")._oleobj_.QueryInterface(kompas6_api5_module.KompasObject.CLSID,
                                                                    pythoncom.IID_IDispatch))

        #  Подключим описание интерфейсов API7
        kompas_api7_module = gencache.EnsureModule("{69AC2981-37C0-4379-84FD-5DD2F3C0A520}", 0, 1, 0)
        application = kompas_api7_module.IApplication(
            Dispatch("Kompas.Application.7")._oleobj_.QueryInterface(kompas_api7_module.IApplication.CLSID,
                                                                    pythoncom.IID_IDispatch))

        Documents = application.Documents
        #  Получим активный документ
        kompas_document = application.ActiveDocument
        kompas_document_3d = kompas_api7_module.IKompasDocument3D(kompas_document)
        iDocument3D = kompas_object.ActiveDocument3D()

        kPart = iDocument3D.GetPart(kompas6_constants_3d.pTop_Part)

        varcoll = kPart.VariableCollection()
        varcoll.refresh()

        values = []

        for i in range(0, 7):
             values.append(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll")) + i).value)

        list_collms = ['N1','L1','N2','L2','a2','N3','L3']

        for i in range(len(values)):
            Variable = varcoll.GetByName(list_collms[i], True, True)
            Variable.value = values[i]

        # Перестраиваем модель
        kPart.RebuildModel()
        # Перерисовываем документ
        iDocument3D.RebuildDocument()

        savePath = kompas_document.PathName[:-4]
        print(savePath)

        kompas_document.SaveAs(savePath + f'{row}.stp')
        kompas_document.Close(True)

    def macroGeomCgange(self,row,geom,sheet_obj,save_path):
        settings = Settings()

        saveCSV = settings.get_from_settings('csv_path')
        saveCSV = saveCSV.replace('"','')
        saveCSV = saveCSV + f'star{row}.csv'
        
        macro_folder = (settings.get_from_settings('macros_path')).replace('\\\\','/')
        macro_folder = macro_folder.replace('"','')
        fout = open(f"{macro_folder}macros{row}.java","w")
        fout.write("""
package macro;

import java.io.File;
import java.io.FileWriter;
import java.io.IOException;
import java.util.*;

import star.common.*;
import star.base.neo.*;
import star.vis.*;

import javax.swing.table.TableModel;

import star.amr.AmrModel;
import star.base.neo.DoubleVector;
import star.base.neo.NeoObjectVector;
import star.base.neo.NeoProperty;
import star.base.neo.StringVector;
import star.cadmodeler.*;
import star.combustion.*;
import star.combustion.fgm.FgmCombustionModel;
import star.combustion.fgm.FgmIdealGasModel;
import star.combustion.fgm.FgmReactionModel;
import star.combustion.fgm.FgmTable;
import star.combustion.tablegenerators.*;
import star.common.*;
import star.dualmesher.VolumeControlDualMesherSizeOption;
import star.emissions.NoxModel;
import star.emissions.NoxThreeEquationZeldovichModel;
import star.emissions.ThermalNoxModel;
import star.flow.MassFlowRateProfile;
import star.flow.VelocityMagnitudeProfile;
import star.keturb.KEpsilonTurbulence;
import star.keturb.KeTwoLayerAllYplusWallTreatment;
import star.keturb.RkeTwoLayerTurbModel;
import star.material.MaterialDataBaseManager;
import star.material.MultiComponentGasModel;
import star.meshing.*;
import star.reactions.ReactingModel;
import star.resurfacer.VolumeControlResurfacerSizeOption;
import star.segregatedenergy.SegregatedFluidEnthalpyModel;
import star.segregatedflow.SegregatedFlowModel;
import star.turbulence.RansTurbulenceModel;
import star.turbulence.TurbulentModel;
import star.vis.*;

import javax.swing.*;

public class %s extends StarMacro {

    public void execute() {

        importGeometry();
        createCylinderParts();
        createVolumeMeshControl();
        createBoundaries();
        generateVolumeMesh();
        saveState();
        createPlaneSection();
        settingPhysicsContinuum();
        createFgmTable();
        createLinePart();
        settingPlaneSection();
        createPlot();
        setStoppingCriterion(%i);
	runSimulation();
        saveCSV();
        saveState();
        Close();
        
    }

    private void importGeometry() {

        Simulation simulation =
                getActiveSimulation();

        CadModel cadModel =
                simulation.get(SolidModelManager.class).createSolidModel();

        cadModel.resetSystemOptions();

        ImportCadFileFeature importCadFileFeature =
                cadModel.importCadFile(resolvePath("%s"), true, false, false, false, false, false, false, true, false, true, NeoProperty.fromString("{\'CGR\': 0, \'IFC\': 0, \'STEP\': 0, \'NX\': 0, \'CATIAV5\': 0, \'SE\': 0, \'JT\': 0}"), false);

        star.cadmodeler.Body cadmodelerBody_0 =
                ((star.cadmodeler.Body) importCadFileFeature.getBodyByIndex(1));

        Face face_0 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{0.01, -0.22786658652431685, 0.10277042815333508})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_0}), "Air blades", false);

        Face face_1 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{0.01, -0.23814856654623856, 0.4957704212230306})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_1}), "Air input", false);

        Face face_2 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{0.01, -0.04266834375631555, -0.13363422057600013})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_2}), "CH4", false);

        Face face_3 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{0.01, -0.042284638094857586, 0.08802686971235965})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_3}), "No Gas 1", false);

        Face face_4 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{0.01, -0.009175941885004056, 0.004994952138007335})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_4}), "No Gas 2", false);

        Face face_5 =
                ((Face) importCadFileFeature.getFaceByLocation(cadmodelerBody_0, new DoubleVector(new double[]{-10.667, 0.6180132, -0.6180132})));

        cadModel.setFaceNameAttributes(new NeoObjectVector(new Object[]{face_5}), "Outlet", false);

        cadmodelerBody_0.getUnNamedFacesDefaultAttributeName();

        simulation.get(SolidModelManager.class).endEditCadModel(cadModel);

        cadModel.createParts(new NeoObjectVector(new Object[]{cadmodelerBody_0}), new NeoObjectVector(new Object[]{}), true, false, 1, false, false, 3, "SharpEdges", 30.0, 2, true, 1.0E-5, false);
    }

    private void createCylinderParts() {

        Simulation simulation =
                getActiveSimulation();

        Units units_0 =
                simulation.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        MeshPartFactory meshPartFactory_0 =
                simulation.get(MeshPartFactory.class);

        SimpleCylinderPart simpleCylinderPart_0 =
                meshPartFactory_0.createNewCylinderPart(simulation.get(SimulationPartManager.class));

        simpleCylinderPart_0.setDoNotRetessellate(true);

        LabCoordinateSystem labCoordinateSystem_0 =
                simulation.getCoordinateSystemManager().getLabCoordinateSystem();

        simpleCylinderPart_0.setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_0.getStartCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_0.getStartCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{0.01, 0.0, 0.0}));

        simpleCylinderPart_0.getEndCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_0.getEndCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-0.667, 0.0, 0.0}));

        simpleCylinderPart_0.getRadius().setUnits(units_0);

        simpleCylinderPart_0.getRadius().setValue(0.721);

        simpleCylinderPart_0.getTessellationDensityOption().setSelected(TessellationDensityOption.Type.MEDIUM);

        simpleCylinderPart_0.rebuildSimpleShapePart();

        simpleCylinderPart_0.setDoNotRetessellate(false);

        SimpleCylinderPart simpleCylinderPart_1 =
                meshPartFactory_0.createNewCylinderPart(simulation.get(SimulationPartManager.class));

        simpleCylinderPart_1.setDoNotRetessellate(true);

        simpleCylinderPart_1.setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_1.getStartCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_1.getStartCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-0.667, 0.0, 0.0}));

        simpleCylinderPart_1.getEndCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_1.getEndCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-1.111, 0.0, 0.0}));

        simpleCylinderPart_1.getRadius().setUnits(units_0);

        simpleCylinderPart_1.getRadius().setValue(0.535);

        simpleCylinderPart_1.getTessellationDensityOption().setSelected(TessellationDensityOption.Type.MEDIUM);

        simpleCylinderPart_1.rebuildSimpleShapePart();

        simpleCylinderPart_1.setDoNotRetessellate(false);

        SimpleCylinderPart simpleCylinderPart_2 =
                meshPartFactory_0.createNewCylinderPart(simulation.get(SimulationPartManager.class));

        simpleCylinderPart_2.setDoNotRetessellate(true);

        simpleCylinderPart_2.setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_2.getStartCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_2.getStartCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-1.111, 0.0, 0.0}));

        simpleCylinderPart_2.getEndCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        simpleCylinderPart_2.getEndCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-10.667, 0.0, 0.0}));

        simpleCylinderPart_2.getRadius().setUnits(units_0);

        simpleCylinderPart_2.getRadius().setValue(1.5);

        simpleCylinderPart_2.getTessellationDensityOption().setSelected(TessellationDensityOption.Type.MEDIUM);

        simpleCylinderPart_2.rebuildSimpleShapePart();

        simpleCylinderPart_2.setDoNotRetessellate(false);
    }

    private void createVolumeMeshControl(){

        Simulation simulation =
                getActiveSimulation();

        Units units_0 =
                simulation.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        SimpleCylinderPart simpleCylinderPart_0 =
                ((SimpleCylinderPart) simulation.get(SimulationPartManager.class).getPart("Cylinder"));

        SimpleCylinderPart simpleCylinderPart_1 =
                ((SimpleCylinderPart) simulation.get(SimulationPartManager.class).getPart("Cylinder 2"));

        SimpleCylinderPart simpleCylinderPart_2 =
                ((SimpleCylinderPart) simulation.get(SimulationPartManager.class).getPart("Cylinder 3"));

        SolidModelPart solidModelPart_0 =
                ((SolidModelPart) simulation.get(SimulationPartManager.class).getPart("Body 1"));

        AutoMeshOperation autoMeshOperation_0 =
                simulation.get(MeshOperationManager.class).createAutoMeshOperation(new StringVector(new String[]{"star.resurfacer.ResurfacerAutoMesher", "star.dualmesher.DualAutoMesher"}), new NeoObjectVector(new Object[]{solidModelPart_0}));

        VolumeCustomMeshControl volumeCustomMeshControl_0 =
                autoMeshOperation_0.getCustomMeshControls().createVolumeControl();

        VolumeCustomMeshControl volumeCustomMeshControl_1 =
                autoMeshOperation_0.getCustomMeshControls().createVolumeControl();

        VolumeCustomMeshControl volumeCustomMeshControl_2 =
                autoMeshOperation_0.getCustomMeshControls().createVolumeControl();

        SurfaceCustomMeshControl surfaceCustomMeshControl_0 =
                autoMeshOperation_0.getCustomMeshControls().createSurfaceControl();

        volumeCustomMeshControl_0.getGeometryObjects().setQuery(null);

        volumeCustomMeshControl_0.getGeometryObjects().setObjects(simpleCylinderPart_0);

        volumeCustomMeshControl_1.getGeometryObjects().setQuery(null);

        volumeCustomMeshControl_1.getGeometryObjects().setObjects(simpleCylinderPart_1);

        volumeCustomMeshControl_2.getGeometryObjects().setQuery(null);

        volumeCustomMeshControl_2.getGeometryObjects().setObjects(simpleCylinderPart_2);

        surfaceCustomMeshControl_0.getGeometryObjects().setQuery(null);

        PartSurface partSurface_0 =
                ((PartSurface) solidModelPart_0.getPartSurfaceManager().getPartSurface("Default"));

        PartSurface partSurface_1 =
                ((PartSurface) solidModelPart_0.getPartSurfaceManager().getPartSurface("Outlet"));

        surfaceCustomMeshControl_0.getGeometryObjects().setObjects(partSurface_0, partSurface_1);

        surfaceCustomMeshControl_0.getCustomConditions().get(PartsTargetSurfaceSizeOption.class).setSelected(PartsTargetSurfaceSizeOption.Type.CUSTOM);

        PartsTargetSurfaceSize partsTargetSurfaceSize_0 =
                surfaceCustomMeshControl_0.getCustomValues().get(PartsTargetSurfaceSize.class);

        Units units_1 =
                ((Units) simulation.getUnitsManager().getObject(""));

        partsTargetSurfaceSize_0.getRelativeSizeScalar().setValueAndUnits(20.0, units_1);

        VolumeControlResurfacerSizeOption volumeControlResurfacerSizeOption_0 =
                volumeCustomMeshControl_0.getCustomConditions().get(VolumeControlResurfacerSizeOption.class);

        volumeControlResurfacerSizeOption_0.setVolumeControlBaseSizeOption(true);

        VolumeControlDualMesherSizeOption volumeControlDualMesherSizeOption_0 =
                volumeCustomMeshControl_0.getCustomConditions().get(VolumeControlDualMesherSizeOption.class);

        volumeControlDualMesherSizeOption_0.setVolumeControlBaseSizeOption(true);

        VolumeControlSize volumeControlSize_0 =
                volumeCustomMeshControl_0.getCustomValues().get(VolumeControlSize.class);

        volumeControlSize_0.getRelativeSizeScalar().setValueAndUnits(2.0, units_1);

        VolumeControlResurfacerSizeOption volumeControlResurfacerSizeOption_1 =
                volumeCustomMeshControl_1.getCustomConditions().get(VolumeControlResurfacerSizeOption.class);

        volumeControlResurfacerSizeOption_1.setVolumeControlBaseSizeOption(true);

        VolumeControlDualMesherSizeOption volumeControlDualMesherSizeOption_1 =
                volumeCustomMeshControl_1.getCustomConditions().get(VolumeControlDualMesherSizeOption.class);

        volumeControlDualMesherSizeOption_1.setVolumeControlBaseSizeOption(true);

        VolumeControlSize volumeControlSize_1 =
                volumeCustomMeshControl_1.getCustomValues().get(VolumeControlSize.class);

        volumeControlSize_1.getRelativeSizeScalar().setValueAndUnits(2.0, units_1);

        VolumeControlResurfacerSizeOption volumeControlResurfacerSizeOption_2 =
                volumeCustomMeshControl_2.getCustomConditions().get(VolumeControlResurfacerSizeOption.class);

        volumeControlResurfacerSizeOption_2.setVolumeControlBaseSizeOption(true);

        VolumeControlDualMesherSizeOption volumeControlDualMesherSizeOption_2 =
                volumeCustomMeshControl_2.getCustomConditions().get(VolumeControlDualMesherSizeOption.class);

        volumeControlDualMesherSizeOption_2.setVolumeControlBaseSizeOption(true);

        VolumeControlSize volumeControlSize_2 =
                volumeCustomMeshControl_2.getCustomValues().get(VolumeControlSize.class);

        volumeControlSize_2.getRelativeSizeScalar().setValueAndUnits(5.0, units_1);

        autoMeshOperation_0.getDefaultValues().get(BaseSize.class).setValueAndUnits(%f, units_0);

        autoMeshOperation_0.getMesherParallelModeOption().setSelected(MesherParallelModeOption.Type.PARALLEL);

    }

    private void createBoundaries() {

        Simulation simulation =
                getActiveSimulation();

        SolidModelPart solidModelPart =
                ((SolidModelPart) simulation.get(SimulationPartManager.class).getPart("Body 1"));

        simulation.getRegionManager().newRegionsFromParts(new NeoObjectVector(new Object[]{solidModelPart}), "OneRegionPerPart", null, "OneBoundaryPerPartSurface", null, "OneFeatureCurve", null, RegionManager.CreateInterfaceMode.BOUNDARY, "OneEdgeBoundaryPerPart", null);

        Region region_0 =
                simulation.getRegionManager().getRegion("Body 1");

        Boundary boundary_0 =
                region_0.getBoundaryManager().getBoundary("Air blades");

        InletBoundary inletBoundary_0 =
                ((InletBoundary) simulation.get(ConditionTypeManager.class).get(InletBoundary.class));

        boundary_0.setBoundaryType(inletBoundary_0);

        Boundary boundary_1 =
                region_0.getBoundaryManager().getBoundary("Air input");

        boundary_1.setBoundaryType(inletBoundary_0);

        Boundary boundary_2 =
                region_0.getBoundaryManager().getBoundary("CH4");

        MassFlowBoundary massFlowBoundary_0 =
                ((MassFlowBoundary) simulation.get(ConditionTypeManager.class).get(MassFlowBoundary.class));

        boundary_2.setBoundaryType(massFlowBoundary_0);

        Boundary boundary_3 =
                region_0.getBoundaryManager().getBoundary("Outlet");

        PressureBoundary pressureBoundary_0 =
                ((PressureBoundary) simulation.get(ConditionTypeManager.class).get(PressureBoundary.class));

        boundary_3.setBoundaryType(pressureBoundary_0);

    }

    private void generateVolumeMesh() {

        Simulation simulation =
                getActiveSimulation();

        MeshPipelineController meshPipelineController =
                simulation.get(MeshPipelineController.class);

        meshPipelineController.generateVolumeMesh();

    }

    private void createPlaneSection() {

        Simulation simulation =
                getActiveSimulation();

        Units units_0 =
                simulation.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        Region region =
                simulation.getRegionManager().getRegion("Body 1");

        PlaneSection planeSection =
                (PlaneSection) simulation.getPartManager().createImplicitPart(new NeoObjectVector(new Object[] {}), new DoubleVector(new double[] {0.0, 0.0, 1.0}), new DoubleVector(new double[] {0.0, 0.0, 0.0}), 0, 1, new DoubleVector(new double[] {0.0}));

        LabCoordinateSystem labCoordinateSystem =
                simulation.getCoordinateSystemManager().getLabCoordinateSystem();

        planeSection.setCoordinateSystem(labCoordinateSystem);

        planeSection.getInputParts().setQuery(null);

        planeSection.getInputParts().setObjects(region);

        planeSection.getOriginCoordinate().setUnits0(units_0);

        planeSection.getOriginCoordinate().setUnits1(units_0);

        planeSection.getOriginCoordinate().setUnits2(units_0);

        planeSection.getOriginCoordinate().setDefinition("");

        planeSection.getOriginCoordinate().setValue(new DoubleVector(new double[] {-5.32849999999975, 0.0, 0.0}));

        planeSection.getOriginCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {-5.32849999999975, 0.0, 0.0}));

        planeSection.getOriginCoordinate().setCoordinateSystem(labCoordinateSystem);

        planeSection.getOrientationCoordinate().setUnits0(units_0);

        planeSection.getOrientationCoordinate().setUnits1(units_0);

        planeSection.getOrientationCoordinate().setUnits2(units_0);

        planeSection.getOrientationCoordinate().setDefinition("");

        planeSection.getOrientationCoordinate().setValue(new DoubleVector(new double[] {0.0, 0.0, 1.0}));

        planeSection.getOrientationCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {0.0, 0.0, 1.0}));

        planeSection.getOrientationCoordinate().setCoordinateSystem(labCoordinateSystem);

        SingleValue singleValue =
                planeSection.getSingleValue();

        singleValue.getValueQuantity().setValue(0.0);

        singleValue.getValueQuantity().setUnits(units_0);

        RangeMultiValue rangeMultiValue =
                planeSection.getRangeMultiValue();

        rangeMultiValue.setNValues(2);

        rangeMultiValue.getStartQuantity().setValue(0.0);

        rangeMultiValue.getStartQuantity().setUnits(units_0);

        rangeMultiValue.getEndQuantity().setValue(1.0);

        rangeMultiValue.getEndQuantity().setUnits(units_0);

        DeltaMultiValue deltaMultiValue =
                planeSection.getDeltaMultiValue();

        deltaMultiValue.setNValues(2);

        deltaMultiValue.getStartQuantity().setValue(0.0);

        deltaMultiValue.getStartQuantity().setUnits(units_0);

        deltaMultiValue.getDeltaQuantity().setValue(1.0);

        deltaMultiValue.getDeltaQuantity().setUnits(units_0);

        MultiValue multiValue =
                planeSection.getArbitraryMultiValue();

        multiValue.getValueQuantities().setUnits(units_0);

        multiValue.getValueQuantities().setArray(new DoubleVector(new double[] {0.0}));

        planeSection.setValueMode(ValueMode.SINGLE);

        PlaneSection planeSection_2 =
                (PlaneSection) simulation.getPartManager().createImplicitPart(new NeoObjectVector(new Object[] {}), new DoubleVector(new double[] {0.0, 0.0, 1.0}), new DoubleVector(new double[] {0.0, 0.0, 0.0}), 0, 1, new DoubleVector(new double[] {0.0}));

        planeSection_2.setCoordinateSystem(labCoordinateSystem);

        planeSection_2.getInputParts().setQuery(null);

        planeSection_2.getInputParts().setObjects(region);

        planeSection_2.getOriginCoordinate().setUnits0(units_0);

    }

    private void settingPhysicsContinuum() {

        Simulation simulation =
                getActiveSimulation();

        PhysicsContinuum physicsContinuum =
                ((PhysicsContinuum) simulation.getContinuumManager().getContinuum("Physics 1"));

        physicsContinuum.enable(MultiComponentGasModel.class);

        physicsContinuum.enable(ReactingModel.class);

        physicsContinuum.enable(FlameletBasedModel.class);

        physicsContinuum.enable(FgmCombustionModel.class);

        physicsContinuum.enable(FgmReactionModel.class);

        physicsContinuum.enable(FgmIdealGasModel.class);

        physicsContinuum.enable(TfcCombustionPartiallyPremixedModel.class);

        physicsContinuum.enable(SegregatedFlowModel.class);

        physicsContinuum.enable(TurbulentModel.class);

        physicsContinuum.enable(SegregatedFluidEnthalpyModel.class);

        physicsContinuum.enable(RansTurbulenceModel.class);

        physicsContinuum.enable(KEpsilonTurbulence.class);

        physicsContinuum.enable(RkeTwoLayerTurbModel.class);

        physicsContinuum.enable(KeTwoLayerAllYplusWallTreatment.class);

        physicsContinuum.enable(SteadyModel.class);

        physicsContinuum.enable(AmrModel.class);

        physicsContinuum.enable(NoxModel.class);

        physicsContinuum.enable(ThermalNoxModel.class);

        physicsContinuum.enable(NoxThreeEquationZeldovichModel.class);

        ProgressVariableIgnitor progressVariableIgnitor =
                physicsContinuum.get(IgnitorManager.class).createIgnitor(ProgressVariableIgnitor.class);

        progressVariableIgnitor.getGeometryPartGroup().setQuery(null);

        SimpleCylinderPart simpleCylinderPart =
                ((SimpleCylinderPart) simulation.get(SimulationPartManager.class).getPart("Cylinder 3"));

        progressVariableIgnitor.getGeometryPartGroup().setObjects(simpleCylinderPart);

        PulseActivator pulseActivator =
                ((PulseActivator) progressVariableIgnitor.getActivator());

        Units units_1 =
                ((Units) simulation.getUnitsManager().getObject(""));

        pulseActivator.getBegin().setValueAndUnits(200.0, units_1);

        pulseActivator.getEnd().setValueAndUnits(300.0, units_1);
    }

    private void createFgmTable() {
        Simulation simulation =
                getActiveSimulation();

        PhysicsContinuum physicsContinuum =
                ((PhysicsContinuum) simulation.getContinuumManager().getContinuum("Physics 1"));

        Region region_0 =
                simulation.getRegionManager().getRegion("Body 1");

        FgmTableGenerator fgmTableGenerator =
                physicsContinuum.get(FgmTableGenerator.class);

        FgmTableParameters fgmTableParameters =
                ((FgmTableParameters) fgmTableGenerator.getTableParameters());

        fgmTableParameters.getTableProgressVariableDefinition().setSelected(TableProgressVariableDefinitionOption.Type.CHEMENTHALPY);

        Fgm0dIgnitionNumericalSettings fgm0dIgnitionNumericalSettings =
                ((Fgm0dIgnitionNumericalSettings) fgmTableParameters.getFgm0dIgnitionNumericalSettings());

        TableAxisParameters tableAxisParameters_0 =
                ((TableAxisParameters) fgm0dIgnitionNumericalSettings.getTableAxisParametersManager().getComponent("Heat Loss Ratio"));

        tableAxisParameters_0.setAdapt(false);

        FixedGridParameters fixedGridParameters_0 =
                tableAxisParameters_0.getFixedGridParameters();

        IntegerValue integerValue_0 =
                fixedGridParameters_0.getDimensionSizeValue();

        integerValue_0.getQuantity().setValue(1.0);

        TableChemistryDefinition tableChemistryDefinition_0 =
                ((TableChemistryDefinition) fgmTableGenerator.getTableChemistryDefinition());

        tableChemistryDefinition_0.importCaseFromChemkin(resolvePath(%s), resolvePath(%s), resolvePath(%s), "", "");

        TableFluidStreamCollection tableFluidStreamCollection_0 =
                ((TableFluidStreamCollection) fgmTableGenerator.getTableFluidStreamCollection());

        TableFluidStream tableFluidStream_0 =
                ((TableFluidStream) tableFluidStreamCollection_0.getOxidizerStream());

        Units units_2 =
                ((Units) simulation.getUnitsManager().getObject("K"));

        tableFluidStream_0.getTemperature().setValueAndUnits(%f, units_2);

        tableFluidStream_0.getFluidStreamComposition().setArray(new DoubleVector(new double[] {0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, %f, 0.0, %f, 0.0, %f, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, %f, 0.0, 0.0, 0.0, 0.0, %f, 0.0, 0.0, 0.0, 0.0, %f, 0.0}));

        TableFluidStream tableFluidStream_1 =
                ((TableFluidStream) tableFluidStreamCollection_0.getFuelStream());

        tableFluidStream_1.getFluidStreamComposition().setArray(new DoubleVector(new double[]{0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0}));

        TableSpeciesForPostProcessing tableSpeciesForPostProcessing_0 =
                ((TableSpeciesForPostProcessing) fgmTableParameters.getTableSpeciesForPostProcessing());

        ((TableSpeciesGroup) tableSpeciesForPostProcessing_0.getTableSpeciesGroup()).setQuery(null);

        star.material.MaterialDataBase materialMaterialDataBase_0 =
                simulation.get(MaterialDataBaseManager.class).getMatlDataBase("Table Generator");

        star.material.DataBaseMaterialManager materialDataBaseMaterialManager_0 =
                materialMaterialDataBase_0.getFolder("Physics 1-Fgm");

        star.material.DataBaseGas materialDataBaseGas_0 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("CH4_Gas"));

        star.material.DataBaseGas materialDataBaseGas_1 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("CO_Gas"));

        star.material.DataBaseGas materialDataBaseGas_2 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("CO2_Gas"));

        star.material.DataBaseGas materialDataBaseGas_3 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("H2O_Gas"));

        star.material.DataBaseGas materialDataBaseGas_4 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("N2_Gas"));

        star.material.DataBaseGas materialDataBaseGas_5 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("NO_Gas"));

        star.material.DataBaseGas materialDataBaseGas_6 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("NO2_Gas"));

        star.material.DataBaseGas materialDataBaseGas_7 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("O2_Gas"));

        star.material.DataBaseGas materialDataBaseGas_8 =
                ((star.material.DataBaseGas) materialDataBaseMaterialManager_0.getMaterial("OH_Gas"));

        ((TableSpeciesGroup) tableSpeciesForPostProcessing_0.getTableSpeciesGroup()).setObjects(materialDataBaseGas_0, materialDataBaseGas_1, materialDataBaseGas_2, materialDataBaseGas_3, materialDataBaseGas_4, materialDataBaseGas_5, materialDataBaseGas_6, materialDataBaseGas_7, materialDataBaseGas_8);

        FgmTable fgmTable_0 =
                ((FgmTable) fgmTableGenerator.getFgmTable());

        fgmTable_0.constructTable();

        Boundary boundary_0 =
                region_0.getBoundaryManager().getBoundary("Air blades");

        VelocityMagnitudeProfile velocityMagnitudeProfile_0 =
                boundary_0.getValues().get(VelocityMagnitudeProfile.class);

        Units units_3 =
                ((Units) simulation.getUnitsManager().getObject("m/s"));

        velocityMagnitudeProfile_0.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_3);

        Boundary boundary_1 =
                region_0.getBoundaryManager().getBoundary("Air input");

        VelocityMagnitudeProfile velocityMagnitudeProfile_1 =
                boundary_1.getValues().get(VelocityMagnitudeProfile.class);

        velocityMagnitudeProfile_1.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_3);

        Boundary boundary_2 =
                region_0.getBoundaryManager().getBoundary("CH4");

        MixtureFractionArrayProfile mixtureFractionArrayProfile_0 =
                boundary_2.getValues().get(MixtureFractionArrayProfile.class);

        mixtureFractionArrayProfile_0.getMethod(ConstantArrayProfileMethod.class).getQuantity().setArray(new DoubleVector(new double[] {1.0}));

        MassFlowRateProfile massFlowRateProfile_0 =
                boundary_2.getValues().get(MassFlowRateProfile.class);

        Units units_4 =
                ((Units) simulation.getUnitsManager().getObject("kg/s"));

        massFlowRateProfile_0.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_4);
    }

    private void createLinePart() {

        Simulation simulation =
                getActiveSimulation();

        LabCoordinateSystem labCoordinateSystem =
                simulation.getCoordinateSystemManager().getLabCoordinateSystem();

        Units units_0 =
                simulation.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        Region region_0 =
                simulation.getRegionManager().getRegion("Body 1");

        LinePart linePart_0 =
                simulation.getPartManager().createLinePart(new NeoObjectVector(new Object[] {}), new DoubleVector(new double[] {0.0, 0.0, 0.0}), new DoubleVector(new double[] {1.0, 0.0, 0.0}), 20);

        linePart_0.getPoint1Coordinate().setCoordinateSystem(labCoordinateSystem);

        linePart_0.getPoint1Coordinate().setUnits0(units_0);

        linePart_0.getPoint1Coordinate().setUnits1(units_0);

        linePart_0.getPoint1Coordinate().setUnits2(units_0);

        linePart_0.getPoint1Coordinate().setDefinition("");

        linePart_0.getPoint1Coordinate().setValue(new DoubleVector(new double[] {-10.666999816894531, 0.0, 0.0}));

        linePart_0.getPoint1Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {-10.666999816894531, 0.0, 0.0}));

        linePart_0.getPoint2Coordinate().setCoordinateSystem(labCoordinateSystem);

        linePart_0.getPoint2Coordinate().setUnits0(units_0);

        linePart_0.getPoint2Coordinate().setUnits1(units_0);

        linePart_0.getPoint2Coordinate().setUnits2(units_0);

        linePart_0.getPoint2Coordinate().setDefinition("");

        linePart_0.getPoint2Coordinate().setValue(new DoubleVector(new double[] {0.009999999776482582, 0.0, 0.0}));

        linePart_0.getPoint2Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {0.009999999776482582, 0.0, 0.0}));

        linePart_0.setCoordinateSystem(labCoordinateSystem);

        linePart_0.getInputParts().setQuery(null);

        linePart_0.getInputParts().setObjects(region_0);

        linePart_0.setResolution(150);

        LinePart linePart_1 =
                simulation.getPartManager().createLinePart(new NeoObjectVector(new Object[] {}), new DoubleVector(new double[] {0.0, 0.0, 0.0}), new DoubleVector(new double[] {1.0, 0.0, 0.0}), 20);

        linePart_1.getPoint1Coordinate().setCoordinateSystem(labCoordinateSystem);

        linePart_1.getPoint1Coordinate().setUnits0(units_0);

        linePart_1.getPoint1Coordinate().setUnits1(units_0);

        linePart_1.getPoint1Coordinate().setUnits2(units_0);

        linePart_1.getPoint1Coordinate().setDefinition("");

        linePart_1.getPoint1Coordinate().setValue(new DoubleVector(new double[] {-3.0, 2.0, 0.0}));

        linePart_1.getPoint1Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {-3.0, 2.0, 0.0}));

        linePart_1.getPoint2Coordinate().setCoordinateSystem(labCoordinateSystem);

        linePart_1.getPoint2Coordinate().setUnits0(units_0);

        linePart_1.getPoint2Coordinate().setUnits1(units_0);

        linePart_1.getPoint2Coordinate().setUnits2(units_0);

        linePart_1.getPoint2Coordinate().setDefinition("");

        linePart_1.getPoint2Coordinate().setValue(new DoubleVector(new double[] {-3.0, -2.0, 0.0}));

        linePart_1.getPoint2Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {-3.0, -2.0, 0.0}));

        linePart_1.setCoordinateSystem(labCoordinateSystem);

        linePart_1.getInputParts().setQuery(null);

        linePart_1.getInputParts().setObjects(region_0);

        linePart_1.setResolution(50);
    }

    private void settingPlaneSection() {

        Simulation simulation =
                getActiveSimulation();

        PlaneSection planeSection_1 =
                ((PlaneSection) simulation.getPartManager().getObject("Plane Section 2"));

        Units units_0 =
                ((Units) simulation.getUnitsManager().getObject("m"));

        planeSection_1.getOriginCoordinate().setUnits1(units_0);

        planeSection_1.getOriginCoordinate().setUnits2(units_0);

        planeSection_1.getOriginCoordinate().setDefinition("");

        planeSection_1.getOriginCoordinate().setValue(new DoubleVector(new double[] {-1.6709601589386591, 0.0, 0.0}));

        planeSection_1.getOriginCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {-1.6709601589386591, 0.0, 0.0}));

        LabCoordinateSystem labCoordinateSystem_0 =
                simulation.getCoordinateSystemManager().getLabCoordinateSystem();

        planeSection_1.getOriginCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        planeSection_1.getOrientationCoordinate().setUnits0(units_0);

        planeSection_1.getOrientationCoordinate().setUnits1(units_0);

        planeSection_1.getOrientationCoordinate().setUnits2(units_0);

        planeSection_1.getOrientationCoordinate().setDefinition("");

        planeSection_1.getOrientationCoordinate().setValue(new DoubleVector(new double[] {1.0, 0.0, 0.0}));

        planeSection_1.getOrientationCoordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[] {1.0, 0.0, 0.0}));

        planeSection_1.getOrientationCoordinate().setCoordinateSystem(labCoordinateSystem_0);

        SingleValue singleValue_1 =
                planeSection_1.getSingleValue();

        singleValue_1.getValueQuantity().setValue(0.0);

        singleValue_1.getValueQuantity().setUnits(units_0);

        RangeMultiValue rangeMultiValue_1 =
                planeSection_1.getRangeMultiValue();

        rangeMultiValue_1.setNValues(2);

        rangeMultiValue_1.getStartQuantity().setValue(0.0);

        rangeMultiValue_1.getStartQuantity().setUnits(units_0);

        rangeMultiValue_1.getEndQuantity().setValue(1.0);

        rangeMultiValue_1.getEndQuantity().setUnits(units_0);

        DeltaMultiValue deltaMultiValue_1 =
                planeSection_1.getDeltaMultiValue();

        deltaMultiValue_1.setNValues(2);

        deltaMultiValue_1.getStartQuantity().setValue(0.0);

        deltaMultiValue_1.getStartQuantity().setUnits(units_0);

        deltaMultiValue_1.getDeltaQuantity().setValue(1.0);

        deltaMultiValue_1.getDeltaQuantity().setUnits(units_0);

        MultiValue multiValue_1 =
                planeSection_1.getArbitraryMultiValue();

        multiValue_1.getValueQuantities().setUnits(units_0);

        multiValue_1.getValueQuantities().setArray(new DoubleVector(new double[] {0.0}));

        planeSection_1.setValueMode(ValueMode.SINGLE);
    }

    private void createPlot() {

        Simulation simulation =
                getActiveSimulation();

        Units units_0 =
                ((Units) simulation.getUnitsManager().getObject("m"));

        XYPlot xYPlot =
                simulation.getPlotManager().createPlot(XYPlot.class);

        PlotUpdate plotUpdate =
                xYPlot.getPlotUpdate();

        HardcopyProperties hardcopyProperties =
                plotUpdate.getHardcopyProperties();

        hardcopyProperties.setCurrentResolutionWidth(25);

        hardcopyProperties.setCurrentResolutionHeight(25);

        xYPlot.getParts().setQuery(null);

        LinePart linePart_0 =
                ((LinePart) simulation.getPartManager().getObject("Line Probe"));

        xYPlot.getParts().setObjects(linePart_0);

        AxisType axisType =
                xYPlot.getXAxisType();

        axisType.getDirectionVector().setComponentsAndUnits(-1.0, 0.0, 0.0, units_0);

        YAxisType yAxisType =
                ((YAxisType) xYPlot.getYAxes().getAxisType("Y Type 1"));

        FieldFunctionUnits fieldFunctionUnits =
                yAxisType.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction =
                ((PrimitiveFieldFunction) simulation.getFieldFunctionManager().getFunction("Temperature"));

        fieldFunctionUnits.setFieldFunction(primitiveFieldFunction);
    }

    private void setStoppingCriterion(int maxStep) {

        Simulation simulation =
                getActiveSimulation();

        StepStoppingCriterion stepStoppingCriterion = ((StepStoppingCriterion) simulation.getSolverStoppingCriterionManager().getSolverStoppingCriterion("Maximum Steps"));
        stepStoppingCriterion.setMaximumNumberSteps(maxStep);

    }

        private void saveState() {

        Simulation simulation =
                getActiveSimulation();

        simulation.saveState("%s");

    }

    private void saveCSV(){
        Simulation simulation_0 =
                getActiveSimulation();

        LinePart linePart_3 =
                simulation_0.getPartManager().createLinePart(new NeoObjectVector(new Object[]{}), new DoubleVector(new double[]{0.0, 0.0, 0.0}), new DoubleVector(new double[]{1.0, 0.0, 0.0}), 20);

        simulation_0.getPartManager().removeObjects(linePart_3);

        Units units_0 =
                simulation_0.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        Region region_0 =
                simulation_0.getRegionManager().getRegion("Body 1");

        LinePart linePart_4 =
                simulation_0.getPartManager().createLinePart(new NeoObjectVector(new Object[]{}), new DoubleVector(new double[]{0.0, 0.0, 0.0}), new DoubleVector(new double[]{1.0, 0.0, 0.0}), 20);

        LabCoordinateSystem labCoordinateSystem_0 =
                simulation_0.getCoordinateSystemManager().getLabCoordinateSystem();

        linePart_4.getPoint1Coordinate().setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getPoint1Coordinate().setUnits0(units_0);

        linePart_4.getPoint1Coordinate().setUnits1(units_0);

        linePart_4.getPoint1Coordinate().setUnits2(units_0);

        linePart_4.getPoint1Coordinate().setDefinition("");

        linePart_4.getPoint1Coordinate().setValue(new DoubleVector(new double[]{-10.667000000000002, 0.0, 0.0}));

        linePart_4.getPoint1Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-10.667000000000002, 0.0, 0.0}));

        linePart_4.getPoint2Coordinate().setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getPoint2Coordinate().setUnits0(units_0);

        linePart_4.getPoint2Coordinate().setUnits1(units_0);

        linePart_4.getPoint2Coordinate().setUnits2(units_0);

        linePart_4.getPoint2Coordinate().setDefinition("");

        linePart_4.getPoint2Coordinate().setValue(new DoubleVector(new double[]{0.010000000000499883, 0.0, 0.0}));

        linePart_4.getPoint2Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{0.010000000000499883, 0.0, 0.0}));

        linePart_4.setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getInputParts().setQuery(null);

        linePart_4.getInputParts().setObjects(region_0);

        linePart_4.setResolution(150);

        XYPlot xYPlot_2 =
                simulation_0.getPlotManager().createPlot(XYPlot.class);

        PlotUpdate plotUpdate_0 =
                xYPlot_2.getPlotUpdate();

        HardcopyProperties hardcopyProperties_1 =
                plotUpdate_0.getHardcopyProperties();

        hardcopyProperties_1.setCurrentResolutionWidth(1277);

        hardcopyProperties_1.setCurrentResolutionHeight(809);

        xYPlot_2.getParts().setQuery(null);

        xYPlot_2.getParts().setObjects(linePart_4);

        AxisType axisType_0 =
                xYPlot_2.getXAxisType();

        axisType_0.getDirectionVector().setComponentsAndUnits(-1.0, 0.0, 0.0, units_0);

        YAxisType yAxisType_0 =
                ((YAxisType) xYPlot_2.getYAxes().getAxisType("Y Type 1"));

        FieldFunctionUnits fieldFunctionUnits_0 =
                yAxisType_0.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_0 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("Temperature"));

        fieldFunctionUnits_0.setFieldFunction(primitiveFieldFunction_0);

        YAxisType yAxisType_1 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_1 =
                yAxisType_1.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_1 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("NitrogenOxide"));

        fieldFunctionUnits_1.setFieldFunction(primitiveFieldFunction_1);

        YAxisType yAxisType_2 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_2 =
                yAxisType_2.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_2 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("MassFractionCO"));

        fieldFunctionUnits_2.setFieldFunction(primitiveFieldFunction_2);

        YAxisType yAxisType_3 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_3 =
                yAxisType_3.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_3 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("MassFractionH2O"));

        fieldFunctionUnits_3.setFieldFunction(primitiveFieldFunction_3);

        Cartesian2DAxisManager cartesian2DAxisManager_0 =
                ((Cartesian2DAxisManager) xYPlot_2.getAxisManager());

        cartesian2DAxisManager_0.setAxesBounds(new Vector(Arrays.<AxisManager.AxisBounds>asList(new AxisManager.AxisBounds("Left Axis", 1.6189140731511915E-16, false, 1922.6049917012276, false), new AxisManager.AxisBounds("Bottom Axis", 0.061179999999502854, false, 10.595819999999998, false))));

        TableModel tm = xYPlot_2.getTableModel();

        try {

            File f = new File(simulation_0.getSessionPath());
            String fileName = %s;
            exportCSV(simulation_0, tm, fileName, ";", ",");

        } catch (IOException e) {
            throw new RuntimeException(e);
        }
    }

    static void exportCSV(Simulation simulation, TableModel tm, String fileName, String delimetr, String doubleDelimetr) throws IOException {
        FileWriter csvWriter = new FileWriter(fileName);

        String header = "";
        for (int i = 0; i < tm.getColumnCount(); i ++) {
            if (i > 0) {
                header = header + delimetr + tm.getColumnName(i);
            } else {
                header = header + tm.getColumnName(i);
            }
        }

        //simulation.println(header);
        csvWriter.append(header);
        csvWriter.append(\"\\n\");

        for (int i = 0; i < tm.getRowCount(); i ++) {
            String data = "";
            for (int j = 0; j < tm.getColumnCount(); j++) {
                if (j > 0) {
                    data = data + ";" + tm.getValueAt(i, j);
                } else {
                    data = data + tm.getValueAt(i, j);
                }
            }
            data = data.replace(".", doubleDelimetr);
            //simulation.println(data);
            csvWriter.append(data);
            csvWriter.append(\"\\n\");
        }

        csvWriter.flush();
        csvWriter.close();
    }
	
        private void runSimulation(){
                Simulation simulation =
                        getActiveSimulation();
                simulation.getSimulationIterator().run();
        }

        private void Close(){
                Simulation simulation =
                        getActiveSimulation();
                simulation.close(ServerConnection.CloseOption.ForceClose);
        }
}       
         """ % (
         f"macros{row}",
         int(settings.get_from_settings('stop_criterion')),
         f"{geom}{row}.stp",
         float(settings.get_from_settings('base_size')),
         settings.get_from_settings('grimech30_path'),
         settings.get_from_settings('thermo30_path'),
         settings.get_from_settings('transport_path'),
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+13).value), #Ar
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+19).value), #C02
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+16).value), #H2
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+20).value), #H2O
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+18).value), #N2
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+15).value), #NH3
         float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+17).value), #O2
         float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+10).value),
         float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+11).value),
         float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+12).value),
         f"{save_path}\\\\star{row}.sim",
         f'"{saveCSV}"'
         ))
        
        fout.close()

    def macroGeomDontChange(self,row,sheet_obj,save_path):
        settings = Settings()

        saveCSV = settings.get_from_settings('csv_path')
        saveCSV = saveCSV.replace('"','')
        saveCSV = saveCSV + f'star{row}.csv'
        
        macro_folder = (settings.get_from_settings('macros_path')).replace('\\\\','/')
        macro_folder = macro_folder.replace('"','')
        fout = open(f"{macro_folder}macros{row}.java","w")
        fout.write("""
package macro;


import java.io.File;
import java.io.FileWriter;
import java.io.IOException;
import java.util.*;

import star.common.*;
import star.base.neo.*;
import star.vis.*;

import javax.swing.table.TableModel;

import star.amr.AmrModel;
import star.base.neo.DoubleVector;
import star.base.neo.NeoObjectVector;
import star.base.neo.NeoProperty;
import star.base.neo.StringVector;
import star.cadmodeler.*;
import star.combustion.*;
import star.combustion.fgm.FgmCombustionModel;
import star.energy.StaticTemperatureProfile;
import star.combustion.fgm.FgmIdealGasModel;
import star.combustion.fgm.FgmReactionModel;
import star.combustion.fgm.FgmTable;
import star.combustion.tablegenerators.*;
import star.common.*;
import star.dualmesher.VolumeControlDualMesherSizeOption;
import star.emissions.NoxModel;
import star.emissions.NoxThreeEquationZeldovichModel;
import star.emissions.ThermalNoxModel;
import star.flow.MassFlowRateProfile;
import star.flow.VelocityMagnitudeProfile;
import star.keturb.KEpsilonTurbulence;
import star.keturb.KeTwoLayerAllYplusWallTreatment;
import star.keturb.RkeTwoLayerTurbModel;
import star.material.MaterialDataBaseManager;
import star.material.MultiComponentGasModel;
import star.meshing.*;
import star.reactions.ReactingModel;
import star.resurfacer.VolumeControlResurfacerSizeOption;
import star.segregatedenergy.SegregatedFluidEnthalpyModel;
import star.segregatedflow.SegregatedFlowModel;
import star.turbulence.RansTurbulenceModel;
import star.turbulence.TurbulentModel;
import star.vis.*;

import javax.swing.*;

public class %s extends StarMacro {

    public void execute() {

        Simulation simulation =
                getActiveSimulation();

        StepStoppingCriterion stepStoppingCriterion = ((StepStoppingCriterion) simulation.getSolverStoppingCriterionManager().getSolverStoppingCriterion("Maximum Steps"));
        stepStoppingCriterion.setMaximumNumberSteps(%i);

        Solution solution =
                simulation.getSolution();

        solution.clearSolution(Solution.Clear.History, Solution.Clear.Fields, Solution.Clear.LagrangianDem);

        Region region =
                simulation.getRegionManager().getRegion("Body 1");

        Units units_0 =
                ((Units) simulation.getUnitsManager().getObject("m/s"));

        Units units_1 =
                ((Units) simulation.getUnitsManager().getObject("kg/s"));

        Units units_2 =
                ((Units) simulation.getUnitsManager().getObject("K"));

        Boundary boundaryAirBlades =
                region.getBoundaryManager().getBoundary("Air blades");

        VelocityMagnitudeProfile velocityMagnitudeProfileAirBlades =
                boundaryAirBlades.getValues().get(VelocityMagnitudeProfile.class);

        velocityMagnitudeProfileAirBlades.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_0);
        boundaryAirBlades.getValues().get(StaticTemperatureProfile.class).getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_2);

        Boundary boundaryAirInlet =
                region.getBoundaryManager().getBoundary("Air input");

        VelocityMagnitudeProfile velocityMagnitudeProfileAirInlet =
                boundaryAirInlet.getValues().get(VelocityMagnitudeProfile.class);

        velocityMagnitudeProfileAirInlet.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_0);
        boundaryAirInlet.getValues().get(StaticTemperatureProfile.class).getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_2);

        Boundary boundaryCH4 =
                region.getBoundaryManager().getBoundary("CH4");

        MassFlowRateProfile massFlowRateProfileCH4 =
                boundaryCH4.getValues().get(MassFlowRateProfile.class);

        massFlowRateProfileCH4.getMethod(ConstantScalarProfileMethod.class).getQuantity().setValueAndUnits(%f, units_1);


        solution.clearSolution(Solution.Clear.History, Solution.Clear.Fields, Solution.Clear.LagrangianDem);

        PhysicsContinuum physicsContinuum =
                ((PhysicsContinuum) simulation.getContinuumManager().getContinuum("Physics 1"));

        FgmTableGenerator fgmTableGenerator =
                physicsContinuum.get(FgmTableGenerator.class);

        FgmTable fgmTable =
                ((FgmTable) fgmTableGenerator.getFgmTable());

        fgmTable.deleteTable();

        TableFluidStreamCollection tableFluidStreamCollection =
                ((TableFluidStreamCollection) fgmTableGenerator.getTableFluidStreamCollection());

        TableFluidStream tableFluidStream =
                ((TableFluidStream) tableFluidStreamCollection.getOxidizerStream());

        tableFluidStream.getFluidStreamComposition().setArray(new DoubleVector(new double[]{0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, %f, 0.0, %f, 0.0, %f, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, %f, 0.0, 0.0, 0.0, 0.0, %f, 0.0, 0.0, 0.0, 0.0, %f, 0.0}));
        //                                                                                  (Ar,   C, C2H, C2H2, C2H3, C2H4, C2H5, C2H6, C3H7, C3H8,  CH, CH2, CH2(S), CH2CHO, CH2CO, CH2O, CH2OH, CH3, CH3CHO, CH3O, CH3OH,    CH4,  CN,  CO,     CO2,   H,    H2, H2CN, H2O, H2O2, HCCO, HCCOH, HCN, HCNN, HCNO, HCO, HNCO, HNO, HO2, HOCN,   N,       N2,      N2O, NCO,  NH, NH2, NH3, NNH,  NO, NO2,   O,       O2,  OH)

        //                tableFluidStream.getTemperature().setValueAndUnits(temperature, units_2);

        TableFluidStream tableFuelStream =
                ((TableFluidStream) tableFluidStreamCollection.getFuelStream());

        tableFuelStream.getFluidStreamComposition().setArray(new DoubleVector(new double[]{0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0}));
        //                                                                                 (Ar,   C, C2H, C2H2, C2H3, C2H4, C2H5, C2H6, C3H7, C3H8,  CH, CH2, CH2(S), CH2CHO, CH2CO, CH2O, CH2OH, CH3, CH3CHO, CH3O, CH3OH,    CH4,  CN,  CO,     CO2,   H,    H2, H2CN, H2O, H2O2, HCCO, HCCOH, HCN, HCNN, HCNO, HCO, HNCO, HNO, HO2, HOCN,   N,       N2,      N2O, NCO,  NH, NH2, NH3, NNH,  NO, NO2,   O,       O2,  OH)


        fgmTable.constructTable();
        simulation.getSimulationIterator().run();
        saveCSV();
        simulation.saveState("%s");
        Close();
    }
    private void saveCSV(){
        Simulation simulation_0 =
                getActiveSimulation();

        LinePart linePart_3 =
                simulation_0.getPartManager().createLinePart(new NeoObjectVector(new Object[]{}), new DoubleVector(new double[]{0.0, 0.0, 0.0}), new DoubleVector(new double[]{1.0, 0.0, 0.0}), 20);

        simulation_0.getPartManager().removeObjects(linePart_3);

        Units units_0 =
                simulation_0.getUnitsManager().getPreferredUnits(Dimensions.Builder().length(1).build());

        Region region_0 =
                simulation_0.getRegionManager().getRegion("Body 1");

        LinePart linePart_4 =
                simulation_0.getPartManager().createLinePart(new NeoObjectVector(new Object[]{}), new DoubleVector(new double[]{0.0, 0.0, 0.0}), new DoubleVector(new double[]{1.0, 0.0, 0.0}), 20);

        LabCoordinateSystem labCoordinateSystem_0 =
                simulation_0.getCoordinateSystemManager().getLabCoordinateSystem();

        linePart_4.getPoint1Coordinate().setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getPoint1Coordinate().setUnits0(units_0);

        linePart_4.getPoint1Coordinate().setUnits1(units_0);

        linePart_4.getPoint1Coordinate().setUnits2(units_0);

        linePart_4.getPoint1Coordinate().setDefinition("");

        linePart_4.getPoint1Coordinate().setValue(new DoubleVector(new double[]{-10.667000000000002, 0.0, 0.0}));

        linePart_4.getPoint1Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{-10.667000000000002, 0.0, 0.0}));

        linePart_4.getPoint2Coordinate().setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getPoint2Coordinate().setUnits0(units_0);

        linePart_4.getPoint2Coordinate().setUnits1(units_0);

        linePart_4.getPoint2Coordinate().setUnits2(units_0);

        linePart_4.getPoint2Coordinate().setDefinition("");

        linePart_4.getPoint2Coordinate().setValue(new DoubleVector(new double[]{0.010000000000499883, 0.0, 0.0}));

        linePart_4.getPoint2Coordinate().setCoordinate(units_0, units_0, units_0, new DoubleVector(new double[]{0.010000000000499883, 0.0, 0.0}));

        linePart_4.setCoordinateSystem(labCoordinateSystem_0);

        linePart_4.getInputParts().setQuery(null);

        linePart_4.getInputParts().setObjects(region_0);

        linePart_4.setResolution(150);

        XYPlot xYPlot_2 =
                simulation_0.getPlotManager().createPlot(XYPlot.class);

        PlotUpdate plotUpdate_0 =
                xYPlot_2.getPlotUpdate();

        HardcopyProperties hardcopyProperties_1 =
                plotUpdate_0.getHardcopyProperties();

        hardcopyProperties_1.setCurrentResolutionWidth(1277);

        hardcopyProperties_1.setCurrentResolutionHeight(809);

        xYPlot_2.getParts().setQuery(null);

        xYPlot_2.getParts().setObjects(linePart_4);

        AxisType axisType_0 =
                xYPlot_2.getXAxisType();

        axisType_0.getDirectionVector().setComponentsAndUnits(-1.0, 0.0, 0.0, units_0);

        YAxisType yAxisType_0 =
                ((YAxisType) xYPlot_2.getYAxes().getAxisType("Y Type 1"));

        FieldFunctionUnits fieldFunctionUnits_0 =
                yAxisType_0.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_0 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("Temperature"));

        fieldFunctionUnits_0.setFieldFunction(primitiveFieldFunction_0);

        YAxisType yAxisType_1 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_1 =
                yAxisType_1.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_1 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("NitrogenOxide"));

        fieldFunctionUnits_1.setFieldFunction(primitiveFieldFunction_1);

        YAxisType yAxisType_2 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_2 =
                yAxisType_2.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_2 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("MassFractionCO"));

        fieldFunctionUnits_2.setFieldFunction(primitiveFieldFunction_2);

        YAxisType yAxisType_3 =
                xYPlot_2.getYAxes().createAxisType();

        FieldFunctionUnits fieldFunctionUnits_3 =
                yAxisType_3.getScalarFunction();

        PrimitiveFieldFunction primitiveFieldFunction_3 =
                ((PrimitiveFieldFunction) simulation_0.getFieldFunctionManager().getFunction("MassFractionH2O"));

        fieldFunctionUnits_3.setFieldFunction(primitiveFieldFunction_3);

        Cartesian2DAxisManager cartesian2DAxisManager_0 =
                ((Cartesian2DAxisManager) xYPlot_2.getAxisManager());

        cartesian2DAxisManager_0.setAxesBounds(new Vector(Arrays.<AxisManager.AxisBounds>asList(new AxisManager.AxisBounds("Left Axis", 1.6189140731511915E-16, false, 1922.6049917012276, false), new AxisManager.AxisBounds("Bottom Axis", 0.061179999999502854, false, 10.595819999999998, false))));

        TableModel tm = xYPlot_2.getTableModel();

        try {

            File f = new File(simulation_0.getSessionPath());
            String fileName = %s;
            exportCSV(simulation_0, tm, fileName, ";", ",");

        } catch (IOException e) {
            throw new RuntimeException(e);
        }
    }

    static void exportCSV(Simulation simulation, TableModel tm, String fileName, String delimetr, String doubleDelimetr) throws IOException {
        FileWriter csvWriter = new FileWriter(fileName);

        String header = "";
        for (int i = 0; i < tm.getColumnCount(); i ++) {
            if (i > 0) {
                header = header + delimetr + tm.getColumnName(i);
            } else {
                header = header + tm.getColumnName(i);
            }
        }

        //simulation.println(header);
        csvWriter.append(header);
        csvWriter.append(\"\\n\");

        for (int i = 0; i < tm.getRowCount(); i ++) {
            String data = "";
            for (int j = 0; j < tm.getColumnCount(); j++) {
                if (j > 0) {
                    data = data + ";" + tm.getValueAt(i, j);
                } else {
                    data = data + tm.getValueAt(i, j);
                }
            }
            data = data.replace(".", doubleDelimetr);
            //simulation.println(data);
            csvWriter.append(data);
            csvWriter.append(\"\\n\");
        }

        csvWriter.flush();
        csvWriter.close();
    }

    private void Close(){
                Simulation simulation =
                        getActiveSimulation();
                simulation.close(ServerConnection.CloseOption.ForceClose);
        }
}        
""" % (f"macros{row}",
        int(settings.get_from_settings('stop_criterion')),
        float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+10).value),
        float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+13).value),
        float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+11).value),
        float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+13).value),
        float(sheet_obj.cell(row=row, column = int(settings.get_from_settings("excel_start_coll"))+12).value),
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+19).value), #C02
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+16).value), #H2
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+20).value), #H2O
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+18).value), #N2
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+15).value), #NH3
        float(sheet_obj.cell(row=row, column=int(settings.get_from_settings("excel_start_coll"))+17).value), #O2
        f"{save_path}\\\\star{row}.sim",
        f'"{saveCSV}"'
        ))

        fout.close()
        
